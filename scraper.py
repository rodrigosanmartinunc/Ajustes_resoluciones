"""
Pipeline: Boletín Oficial CBA → Filtro ERSeP → Excel
Busca PDFs del boletín, filtra resoluciones de ERSeP y extrae campos clave.
"""

import requests
import re
import json
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

import pdfplumber
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

# ── Configuración ────────────────────────────────────────────────────────────

BASE_URL = "https://boletinoficial.cba.gov.ar"
PDF_BASE = f"{BASE_URL}/wp-content/4p96humuzp"
OUTPUT_EXCEL = "resultados_ersep.xlsx"
PDF_DIR = Path("pdfs_descargados")
PDF_DIR.mkdir(exist_ok=True)

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-AR,es;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
}

# Palabras clave para filtrar resoluciones relevantes
PALABRAS_CLAVE = ["ERSeP", "Resolución General", "cooperativa", "COOPERATIVA", "tarifario", "agua corriente"]

# Campos a extraer con sus patrones regex
PATRONES = {
    "resolucion":    r"Resolución\s+(?:General\s+)?N[°º\.]\s*(\d+[\-/]?\d*)",
    "anio":          r"\b(20\d{2})\b",
    "periodo_costos": r"[Pp]eríodo\s+(?:de\s+)?[Cc]ostos[:\s]+([A-Za-z]+\s+\d{4}\s*[-–]\s*[A-Za-z]+\s+\d{4}|\w+\s+\d{4})",
    "periodo_inicio": r"[Pp]eríodo\s+[Ii]nicio[:\s]+([\d/\-\.]+)",
    "periodo_cierre": r"[Pp]eríodo\s+[Cc]ierre[:\s]+([\d/\-\.]+)",
    "porcentaje":    r"(\d{1,3}[,\.]\d{1,2})\s*%",
    "rige_desde":    r"[Rr]ige\s+[Dd]esde[:\s]+([\d/\-\.]+)",
    "prestadora":    r"(?:COOPERATIVA[^\.]{5,80}LTDA\.?|cooperativa[^\.]{5,80}ltda\.?)",
    "cuit":          r"CUIT[:\s#Nº°]*(\d{2}-\d{8}-\d{1}|\d{11})",
    "expediente":    r"[Ee]xp(?:ediente)?[:\s\.#Nº°]*([A-Z0-9\-/]{5,30})",
    "tarifa_agua":   r"[Tt]arifa.*?[Aa]gua.*?[\$\s]([\d\.,]+)",
    "tarifa_cloacas": r"[Tt]arifa.*?[Cc]loaca.*?[\$\s]([\d\.,]+)",
}

# ── Funciones de scraping ────────────────────────────────────────────────────

def get_session():
    s = requests.Session()
    s.headers.update(HEADERS)
    return s


def obtener_urls_pdfs(session, dias_atras=30):
    """Obtiene URLs de PDFs del boletín desde la página principal (últimos N días)."""
    urls = []
    try:
        resp = session.get(f"{BASE_URL}/", timeout=30)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")

        # Buscar todos los links a PDFs en la página
        for a in soup.find_all("a", href=True):
            href = a["href"]
            if ".pdf" in href.lower() and "wp-content" in href:
                if not href.startswith("http"):
                    href = BASE_URL + href
                urls.append(href)

        # Si no hay links directos, construir URLs por fecha
        if not urls:
            urls = construir_urls_por_fecha(dias_atras)

    except Exception as e:
        print(f"[WARN] No se pudo parsear la home: {e}. Construyendo URLs por fecha...")
        urls = construir_urls_por_fecha(dias_atras)

    # Deduplicar
    urls = list(dict.fromkeys(urls))
    print(f"[INFO] {len(urls)} PDFs encontrados")
    return urls


def construir_urls_por_fecha(dias_atras=30):
    """
    Construye posibles URLs de PDFs basándose en el patrón observado:
    /wp-content/4p96humuzp/YYYY/MM/NOMBRE.pdf
    Las secciones típicas son: 1_Secc, 2_Secc, 3_Secc, 4_Secc, 5_Secc
    con formato de fecha DDMMYY al final.
    """
    urls = []
    secciones = ["1_Secc", "2_Secc", "3_Secc", "4_Secc", "5_Secc"]
    hoy = datetime.today()

    for i in range(dias_atras):
        fecha = hoy - timedelta(days=i)
        # Saltar fines de semana (el boletín no se publica sábados/domingos)
        if fecha.weekday() >= 5:
            continue
        yyyy = fecha.strftime("%Y")
        mm = fecha.strftime("%m")
        dd_mm_yy = fecha.strftime("%d%m%y")

        for sec in secciones:
            url = f"{PDF_BASE}/{yyyy}/{mm}/{sec}_{dd_mm_yy}.pdf"
            urls.append(url)

    return urls


def descargar_pdf(session, url, destino):
    """Descarga un PDF y lo guarda en disco. Retorna True si tuvo éxito."""
    try:
        resp = session.get(url, timeout=30, stream=True)
        if resp.status_code == 200 and "pdf" in resp.headers.get("Content-Type", "").lower():
            with open(destino, "wb") as f:
                for chunk in resp.iter_content(8192):
                    f.write(chunk)
            return True
    except Exception as e:
        pass
    return False


# ── Funciones de extracción ──────────────────────────────────────────────────

def contiene_ersep(texto):
    """Devuelve True si el texto contiene alguna palabra clave relevante."""
    return any(kw in texto for kw in PALABRAS_CLAVE)


def extraer_campos(texto, url_pdf, fecha_boletin):
    """Extrae campos clave del texto usando regex."""
    campos = {
        "url_pdf": url_pdf,
        "fecha_boletin": fecha_boletin,
        "anio": None,
        "resolucion": None,
        "periodo_costos": None,
        "periodo_inicio": None,
        "periodo_cierre": None,
        "porcentaje_otorgado": None,
        "rige_desde": None,
        "prestadora": None,
        "cuit": None,
        "expediente": None,
        "tarifa_agua": None,
        "tarifa_cloacas": None,
        "texto_fragmento": texto[:500].replace("\n", " "),
    }

    for campo, patron in PATRONES.items():
        m = re.search(patron, texto, re.IGNORECASE)
        if m:
            campos[campo] = m.group(1).strip() if m.lastindex else m.group(0).strip()

    return campos


def procesar_pdf(pdf_path, url_pdf):
    """
    Abre un PDF, busca páginas con contenido ERSeP y extrae datos.
    Retorna lista de dicts (una entrada por resolución encontrada).
    """
    resultados = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Intentar detectar la fecha del boletín desde el nombre del archivo
            nombre = Path(pdf_path).stem
            fecha_boletin = extraer_fecha_nombre(nombre)

            for num_pag, pagina in enumerate(pdf.pages, 1):
                texto = pagina.extract_text() or ""
                if contiene_ersep(texto):
                    campos = extraer_campos(texto, url_pdf, fecha_boletin)
                    campos["pagina"] = num_pag
                    resultados.append(campos)
                    print(f"  [✓] ERSeP encontrado en pág {num_pag} — {campos.get('prestadora', 'prestadora no detectada')}")

    except Exception as e:
        print(f"  [ERR] No se pudo procesar {pdf_path}: {e}")

    return resultados


def extraer_fecha_nombre(nombre_archivo):
    """Intenta extraer la fecha del nombre del archivo (ej: 1_Secc_180326 → 2026-03-18)."""
    m = re.search(r"(\d{2})(\d{2})(\d{2})$", nombre_archivo)
    if m:
        dd, mm, yy = m.groups()
        anio = f"20{yy}"
        try:
            return datetime(int(anio), int(mm), int(dd)).strftime("%Y-%m-%d")
        except:
            pass
    return datetime.today().strftime("%Y-%m-%d")


# ── Generación del Excel ─────────────────────────────────────────────────────

COLUMNAS = [
    ("Fecha Boletín",       "fecha_boletin",        18),
    ("Año",                 "anio",                  8),
    ("Resolución",          "resolucion",            20),
    ("Período Costos",      "periodo_costos",        25),
    ("Período Inicio",      "periodo_inicio",        16),
    ("Período Cierre",      "periodo_cierre",        16),
    ("% Otorgado",          "porcentaje_otorgado",   14),
    ("Rige Desde",          "rige_desde",            16),
    ("Prestadora",          "prestadora",            45),
    ("CUIT",                "cuit",                  18),
    ("Expediente",          "expediente",            20),
    ("Tarifa Agua",         "tarifa_agua",           14),
    ("Tarifa Cloacas",      "tarifa_cloacas",        14),
    ("Página PDF",          "pagina",                10),
    ("URL PDF",             "url_pdf",               60),
    ("Fragmento Texto",     "texto_fragmento",       60),
]

COLOR_HEADER   = "1F4E79"   # azul oscuro
COLOR_SUBHEAD  = "D6E4F0"   # azul claro para filas pares
COLOR_ACCENT   = "F0B429"   # amarillo para % otorgado


def crear_excel(registros, output_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Resoluciones ERSeP"
    ws.freeze_panes = "A2"

    # ── Header ──
    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color=COLOR_HEADER)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border  = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_idx, (titulo, _, ancho) in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=1, column=col_idx, value=titulo)
        cell.font   = header_font
        cell.fill   = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho

    ws.row_dimensions[1].height = 32

    # ── Datos ──
    fill_par  = PatternFill("solid", start_color=COLOR_SUBHEAD)
    fill_impar = PatternFill("solid", start_color="FFFFFF")
    fill_pct  = PatternFill("solid", start_color="FFF3CD")
    data_font  = Font(name="Arial", size=10)
    data_align = Alignment(vertical="center", wrap_text=False)

    col_pct = next((i+1 for i, (_, k, _) in enumerate(COLUMNAS) if k == "porcentaje_otorgado"), None)

    for row_idx, reg in enumerate(registros, 2):
        fill_base = fill_par if row_idx % 2 == 0 else fill_impar
        for col_idx, (_, key, _) in enumerate(COLUMNAS, 1):
            valor = reg.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=valor or "")
            cell.font      = data_font
            cell.alignment = data_align
            cell.border    = thin_border
            # Resaltar la columna de porcentaje
            if col_idx == col_pct and valor:
                cell.fill = fill_pct
                cell.font = Font(name="Arial", size=10, bold=True)
            else:
                cell.fill = fill_base
        ws.row_dimensions[row_idx].height = 18

    # ── Hoja de log / metadata ──
    ws_log = wb.create_sheet("Log Ejecución")
    ws_log["A1"] = "Pipeline ERSeP — Log de ejecución"
    ws_log["A1"].font = Font(name="Arial", bold=True, size=12)
    ws_log["A3"] = "Fecha ejecución"
    ws_log["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws_log["A4"] = "Total registros encontrados"
    ws_log["B4"] = len(registros)
    ws_log["A5"] = "Palabras clave usadas"
    ws_log["B5"] = ", ".join(PALABRAS_CLAVE)
    for col in ["A", "B"]:
        ws_log.column_dimensions[col].width = 35

    wb.save(output_path)
    print(f"[✓] Excel guardado: {output_path} ({len(registros)} registros)")


# ── Orquestador principal ────────────────────────────────────────────────────

def main(dias_atras=30):
    print(f"=== Pipeline ERSeP — {datetime.now().strftime('%Y-%m-%d %H:%M')} ===")
    session = get_session()

    # 1. Obtener URLs de PDFs
    urls_pdf = obtener_urls_pdfs(session, dias_atras=dias_atras)

    # 2. Descargar y procesar cada PDF
    todos_registros = []
    for url in urls_pdf:
        nombre_archivo = url.split("/")[-1]
        destino = PDF_DIR / nombre_archivo

        # Descargar si no existe ya
        if not destino.exists():
            ok = descargar_pdf(session, url, destino)
            if not ok:
                continue  # PDF no disponible para esa fecha
        
        print(f"[→] Procesando: {nombre_archivo}")
        registros = procesar_pdf(destino, url)
        todos_registros.extend(registros)

    # 3. Generar Excel
    if todos_registros:
        crear_excel(todos_registros, OUTPUT_EXCEL)
    else:
        print("[INFO] No se encontraron resoluciones ERSeP en el período consultado.")
        # Crear Excel vacío con encabezados para que el artifact no falle
        crear_excel([], OUTPUT_EXCEL)

    # 4. Guardar resumen JSON (para debugging en Actions)
    resumen = {
        "fecha_ejecucion": datetime.now().isoformat(),
        "pdfs_procesados": len(urls_pdf),
        "registros_encontrados": len(todos_registros),
        "registros": todos_registros[:5],  # preview de los primeros 5
    }
    with open("resumen.json", "w", encoding="utf-8") as f:
        json.dump(resumen, f, ensure_ascii=False, indent=2)

    print(f"=== Fin. {len(todos_registros)} registros ERSeP encontrados. ===")
    return todos_registros


if __name__ == "__main__":
    dias = int(sys.argv[1]) if len(sys.argv) > 1 else 30
    main(dias_atras=dias)
